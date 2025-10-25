"""
Excel export functionality with charts and analytics.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from .models import WellbeingRecord
from .analytics_utils import (
    UserSegmentation, CorrelationAnalysis, PredictiveTrends, HeatmapData
)
from django.contrib.auth import get_user_model
User = get_user_model()
from django.db.models import Avg 

class WellbeingExcelExport:
    """Generate comprehensive Excel reports with charts."""
    
    def __init__(self, days=30):
        self.days = days
        self.start_date = timezone.now().date() - timedelta(days=days)
        self.records = WellbeingRecord.objects.filter(date__gte=self.start_date)
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        
        # Define styles
        self.header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_summary_sheet(self):
        """Create summary statistics sheet."""
        ws = self.wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "Wellbeing Analytics Report"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Period: {self.start_date} to {timezone.now().date()}"
        ws['A2'].font = Font(italic=True, size=10)
        ws.merge_cells('A2:D2')
        
        # Statistics
        row = 4
        stats = {
            'Total Records': self.records.count(),
            'Active Users': self.records.values('user').distinct().count(),
            'Avg Mood Score': round(self.records.aggregate(Avg('mood_score'))['mood_score__avg'] or 0, 2),
            'Avg Energy Level': round(self.records.aggregate(Avg('energy_level'))['energy_level__avg'] or 0, 2),
            'Avg Sleep Hours': round(self.records.aggregate(Avg('sleep_hours'))['sleep_hours__avg'] or 0, 2),
            'Avg Productivity': round(self.records.aggregate(Avg('productivity_score'))['productivity_score__avg'] or 0, 2),
        }
        
        for label, value in stats.items():
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def create_records_sheet(self):
        """Create detailed records sheet."""
        ws = self.wb.create_sheet("Records")
        
        # Headers
        headers = ['Date', 'User', 'Mood Score', 'Energy Level', 'Sleep Hours', 'Productivity', 'AI Summary']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data
        records = self.records.select_related('user').order_by('-date')
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1).value = record.date
            ws.cell(row=row, column=2).value = record.user.username
            ws.cell(row=row, column=3).value = record.mood_score
            ws.cell(row=row, column=4).value = record.energy_level
            ws.cell(row=row, column=5).value = record.sleep_hours
            ws.cell(row=row, column=6).value = record.productivity_score
            ws.cell(row=row, column=7).value = record.ai_summary[:50] if record.ai_summary else ""
            
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self.border
        
        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 30
    
    def create_segmentation_sheet(self):
        """Create user segmentation sheet."""
        ws = self.wb.create_sheet("Segmentation")
        
        # Title
        ws['A1'] = "User Segmentation"
        ws['A1'].font = self.title_font
        
        # Get segments
        segments = UserSegmentation.get_all_user_segments(self.days)
        
        row = 3
        for segment, users in segments.items():
            ws[f'A{row}'] = segment
            ws[f'A{row}'].font = Font(bold=True, size=11)
            ws[f'B{row}'] = len(users)
            ws[f'B{row}'].font = Font(bold=True, size=11)
            row += 1
            
            for user_info in users:
                ws[f'A{row}'] = user_info['username']
                ws[f'B{row}'] = user_info['email']
                row += 1
            
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def create_correlations_sheet(self):
        """Create correlations sheet."""
        ws = self.wb.create_sheet("Correlations")
        
        # Title
        ws['A1'] = "Variable Correlations"
        ws['A1'].font = self.title_font
        
        # Calculate correlations
        records_list = list(self.records.order_by('date'))
        correlations = CorrelationAnalysis.calculate_correlations(records_list)
        insights = CorrelationAnalysis.get_correlation_insights(correlations)
        
        # Headers
        headers = ['Variable Pair', 'Correlation', 'Strength', 'Direction']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Data
        for row, insight in enumerate(insights, 4):
            ws.cell(row=row, column=1).value = insight['metric']
            ws.cell(row=row, column=2).value = insight['correlation']
            ws.cell(row=row, column=3).value = insight['strength']
            ws.cell(row=row, column=4).value = insight['direction']
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self.border
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def create_predictions_sheet(self):
        """Create at-risk users sheet."""
        ws = self.wb.create_sheet("Predictions")
        
        # Title
        ws['A1'] = "Users at Risk (Declining Trends)"
        ws['A1'].font = self.title_font
        
        # Get at-risk users
        at_risk_users = PredictiveTrends.identify_at_risk_users(self.days)
        
        # Headers
        headers = ['Username', 'Trend', 'Latest Mood', 'Risk Level']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Data
        for row, user_risk in enumerate(at_risk_users, 4):
            ws.cell(row=row, column=1).value = user_risk['username']
            ws.cell(row=row, column=2).value = user_risk['trend']
            ws.cell(row=row, column=3).value = user_risk['latest_mood']
            ws.cell(row=row, column=4).value = user_risk['risk_level']
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self.border
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
    
    def create_heatmap_sheet(self):
        """Create heatmap data sheet."""
        ws = self.wb.create_sheet("Heatmap")
        
        # Title
        ws['A1'] = "Mood by Day of Week"
        ws['A1'].font = self.title_font
        
        # Get heatmap data
        heatmap = HeatmapData.generate_daily_heatmap(self.days)
        
        # Headers
        headers = ['Day', 'Avg Mood', 'Record Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Data
        for row, day_data in enumerate(heatmap, 4):
            ws.cell(row=row, column=1).value = day_data['day']
            ws.cell(row=row, column=2).value = day_data['mood']
            ws.cell(row=row, column=3).value = day_data['count']
            
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.border
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
    
    def generate(self):
        """Generate complete Excel workbook."""
        self.create_summary_sheet()
        self.create_records_sheet()
        self.create_segmentation_sheet()
        self.create_correlations_sheet()
        self.create_predictions_sheet()
        self.create_heatmap_sheet()
        
        return self.wb


def export_wellbeing_excel(request):
    """View to export wellbeing data to Excel."""
    from django.contrib.admin.views.decorators import staff_member_required
    from django.db.models import Avg
    
    days = int(request.GET.get('days', 30))
    
    # Generate workbook
    exporter = WellbeingExcelExport(days)
    wb = exporter.generate()
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="wellbeing_report_{timezone.now().date()}.xlsx"'
    
    wb.save(response)
    return response
